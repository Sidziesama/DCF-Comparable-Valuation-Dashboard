"""Professional, reusable Excel export for public-company valuation models."""

from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Mapping

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


NAVY, BLUE, LIGHT_BLUE = "17365D", "0000FF", "D9EAF7"
WHITE, BLACK, GREEN = "FFFFFF", "000000", "008000"
GRAY, LIGHT_GRAY, RED = "D9E1F2", "F2F2F2", "C00000"
INPUT_FILL, PASS_FILL, FAIL_FILL = "FFF2CC", "E2F0D9", "FCE4D6"
MONEY = '$#,##0;[Red]($#,##0);-'
PER_SHARE = '$0.00;[Red]($0.00);-'
PERCENT = '0.0%;[Red](0.0%);-'
MULTIPLE = '0.0x;[Red](0.0x);-'
NUMBER = '#,##0.0;[Red](#,##0.0);-'


def _safe(value):
    if pd.isna(value):
        return None
    if isinstance(value, np.generic):
        return value.item()
    return value


def _title(ws, title, subtitle=None, end_col=8):
    ws.sheet_view.showGridLines = False
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    cell = ws.cell(1, 1, title)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(name="Arial", size=15, bold=True, color=WHITE)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 27
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
        ws.cell(2, 1, subtitle).font = Font(name="Arial", size=9, italic=True, color="666666")


def _section(ws, row, label, end_col=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    cell = ws.cell(row, 1, label)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.alignment = Alignment(horizontal="left")
    return row + 1


def _write_frame(ws, frame, start_row=4, start_col=1, *, include_index=True):
    data = frame.copy()
    if include_index:
        index_name = data.index.name or "Period"
        data.insert(0, index_name, data.index)
    headers = list(data.columns)
    for offset, header in enumerate(headers):
        cell = ws.cell(start_row, start_col + offset, str(header).replace("_", " ").title())
        cell.fill = PatternFill("solid", fgColor=GRAY)
        cell.font = Font(name="Arial", size=9, bold=True)
        cell.alignment = Alignment(horizontal="right" if offset else "left")
        cell.border = Border(bottom=Side(style="thin", color=NAVY))
    for r_offset, row in enumerate(data.itertuples(index=False, name=None), 1):
        for c_offset, value in enumerate(row):
            cell = ws.cell(start_row + r_offset, start_col + c_offset, _safe(value))
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(horizontal="right" if c_offset else "left")
            name = str(headers[c_offset]).lower()
            if any(token in name for token in ("margin", "growth", "rate", "weight", "pct", "upside", "assumption", "bound", "wacc")):
                cell.number_format = PERCENT
            elif "multiple" in name:
                cell.number_format = MULTIPLE
            elif "price" in name:
                cell.number_format = PER_SHARE
            elif name in ("period", "year") and isinstance(value, (int, float, np.number)):
                cell.number_format = "0"
            elif isinstance(value, (int, float, np.number)):
                cell.number_format = NUMBER
    ws.freeze_panes = ws.cell(start_row + 1, start_col + (1 if include_index else 0))
    return start_row + len(data) + 1, start_col + len(headers) - 1


def _polish(ws):
    used_col = ws.max_column
    for column in range(1, used_col + 1):
        letter = get_column_letter(column)
        max_len = max((len(str(ws.cell(row, column).value or "")) for row in range(1, ws.max_row + 1)), default=0)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 11), 34)
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width, 26)
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and not cell.font.name:
                cell.font = Font(name="Arial", size=9)


def _assumption_value(assumptions, key, default=None):
    value = assumptions.get(key, default)
    return float(value) if isinstance(value, (int, float, np.number)) else value


def export_valuation_workbook(
    output_path,
    *,
    company_name,
    ticker,
    historical,
    forecasts,
    assumptions,
    wacc,
    terminal_growth,
    cash,
    debt,
    shares_outstanding,
    current_price,
    reverse_dcf=None,
    statements=None,
    trading_comps=None,
    dcf_sensitivity=None,
    football_field=None,
    model_checks=None,
    analytics=None,
    wacc_report=None,
):
    """Create an analyst-style workbook and return the saved path."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    names = ["Summary", "Assumptions", "Historical Financials", "Forecast & 3 Statements",
             "Operating Drivers", "WACC", "DCF", "Reverse DCF", "Trading Comps",
             "Sensitivities", "Football Field", "Model Checks & Analytics"]
    for name in names:
        wb.create_sheet(name)

    # Assumptions: visible source of truth for valuation formulas.
    ws = wb["Assumptions"]
    _title(ws, "Key Model Assumptions", "Blue font / yellow fill = editable input", 5)
    rows = [
        ("Company", company_name, "Text"), ("Ticker", ticker, "Text"),
        ("Valuation date", date.today(), "Date"), ("Current share price", current_price, "$/share"),
        ("Cash", cash, "$mm"), ("Debt", debt, "$mm"),
        ("Shares outstanding", shares_outstanding, "mm"), ("WACC", wacc, "%"),
        ("Terminal growth", terminal_growth, "%"),
    ]
    ws.append([]); ws.append(["Assumption", "Value", "Units", "Notes"])
    for cell in ws[4]:
        cell.fill = PatternFill("solid", fgColor=GRAY); cell.font = Font(bold=True, name="Arial")
    for row_idx, (label, value, unit) in enumerate(rows, 5):
        ws.cell(row_idx, 1, label); input_cell = ws.cell(row_idx, 2, value); ws.cell(row_idx, 3, unit)
        input_cell.font = Font(name="Arial", color=BLUE); input_cell.fill = PatternFill("solid", fgColor=INPUT_FILL)
        input_cell.number_format = ("mmm d, yyyy" if unit == "Date" else
                                    PERCENT if unit == "%" else
                                    PER_SHARE if unit == "$/share" else NUMBER)
    assumption_cells = {label: row for row, (label, _, _) in enumerate(rows, 5)}
    ws.freeze_panes = "B5"

    # Historical and operating forecasts.
    ws = wb["Historical Financials"]
    _title(ws, f"{company_name} Historical Financials", "Reported / LTM financial profile ($mm except per-share data)", 12)
    _write_frame(ws, historical, 4, 1)

    ws = wb["Operating Drivers"]
    _title(ws, "Operating Drivers", "Bear / Base / Bull explicit forecast assumptions", 12)
    row = 4
    for scenario in ("bear", "base", "bull"):
        if scenario not in forecasts: continue
        row = _section(ws, row, scenario.upper(), 12)
        columns = [c for c in ("revenue", "revenue_growth", "operating_margin", "tax_rate", "fcff_margin") if c in forecasts[scenario]]
        row, _ = _write_frame(ws, forecasts[scenario][columns], row, 1)
        row += 2

    ws = wb["Forecast & 3 Statements"]
    _title(ws, "Forecast & Three Statements", "Base-case linked statements; scenario details remain available in Operating Drivers", 18)
    row = 4
    if statements and "base" in statements:
        for label, key in (("Income Statement", "income_statement"), ("Balance Sheet", "balance_sheet"),
                           ("Cash Flow Statement", "cash_flow_statement"), ("FCFF Bridge", "fcff_forecast"),
                           ("Working Capital Schedule", "working_capital_schedule"),
                           ("PP&E / Depreciation Schedule", "ppe_schedule"),
                           ("Debt / Interest Schedule", "debt_schedule"),
                           ("Equity / Share Count Schedule", "equity_schedule"),
                           ("Capital Returns Schedule", "capital_returns_schedule")):
            if key in statements["base"]:
                row = _section(ws, row, label, 18)
                row, _ = _write_frame(ws, statements["base"][key], row, 1)
                row += 2
    else:
        _write_frame(ws, forecasts["base"], 4, 1)

    # WACC.
    ws = wb["WACC"]
    _title(ws, "Weighted Average Cost of Capital", "Capital structure and cost-of-capital build", 8)
    if wacc_report is not None and not wacc_report.empty:
        _write_frame(ws, wacc_report, 4, 1, include_index=False)
    else:
        ws["A4"], ws["B4"] = "WACC", f"='Assumptions'!B{assumption_cells['WACC']}"
        ws["B4"].number_format = PERCENT; ws["B4"].font = Font(color=GREEN)

    # Formula-driven DCF linked to the base forecast values and assumption cells.
    ws = wb["DCF"]
    _title(ws, "Discounted Cash Flow Valuation", "Gordon growth method; forecast cash flows linked to model outputs", 12)
    headers = ["Year", "Revenue", "EBIT", "D&A", "Capex", "Change in NWC", "FCFF", "Period", "Discount Factor", "PV of FCFF"]
    for col, header in enumerate(headers, 1):
        c = ws.cell(4, col, header); c.fill = PatternFill("solid", fgColor=GRAY); c.font = Font(bold=True, name="Arial")
    base = forecasts["base"]
    for idx, (period, values) in enumerate(base.iterrows(), 5):
        ws.cell(idx, 1, _safe(period))
        for col, field in enumerate(("revenue", "ebit", "da", "capex", "change_nwc", "fcff"), 2):
            ws.cell(idx, col, _safe(values.get(field)))
        ws.cell(idx, 8, idx - 4)
        ws.cell(idx, 9, f"=1/(1+'Assumptions'!B{assumption_cells['WACC']})^H{idx}")
        ws.cell(idx, 10, f"=G{idx}*I{idx}")
        for col in range(2, 11): ws.cell(idx, col).number_format = NUMBER
        ws.cell(idx, 9).number_format = PERCENT
    terminal_row = 5 + len(base)
    last = terminal_row - 1
    labels = ["PV of forecast FCFF", "Terminal value", "PV of terminal value", "Enterprise value", "Less: debt", "Add: cash", "Equity value", "Shares outstanding", "Implied share price"]
    formulas = [f"=SUM(J5:J{last})", f"=G{last}*(1+'Assumptions'!B{assumption_cells['Terminal growth']})/('Assumptions'!B{assumption_cells['WACC']}-'Assumptions'!B{assumption_cells['Terminal growth']})",
                f"=B{terminal_row+1}*I{last}", f"=SUM(B{terminal_row},B{terminal_row+2})", f"='Assumptions'!B{assumption_cells['Debt']}",
                f"='Assumptions'!B{assumption_cells['Cash']}", f"=B{terminal_row+3}-B{terminal_row+4}+B{terminal_row+5}",
                f"='Assumptions'!B{assumption_cells['Shares outstanding']}", f"=B{terminal_row+6}/B{terminal_row+7}"]
    for offset, (label, formula) in enumerate(zip(labels, formulas)):
        r = terminal_row + offset; ws.cell(r, 1, label); ws.cell(r, 2, formula)
        ws.cell(r, 2).font = Font(name="Arial", color=GREEN); ws.cell(r, 2).number_format = PER_SHARE if label == "Implied share price" else NUMBER
    ws.cell(terminal_row + 8, 1).border = Border(top=Side(style="double", color=NAVY))
    ws.cell(terminal_row + 8, 2).border = Border(top=Side(style="double", color=NAVY))
    ws.freeze_panes = "B5"

    # Reverse DCF and other output tabs.
    ws = wb["Reverse DCF"]
    _title(ws, "Reverse DCF", "Market-implied assumptions required to reconcile to the current share price", 11)
    if reverse_dcf is not None and not reverse_dcf.empty:
        _write_frame(ws, reverse_dcf, 4, 1)
    else:
        ws["A4"] = "Reverse DCF output unavailable"

    for sheet_name, frame, subtitle in (
        ("Trading Comps", trading_comps, "Selected peer trading multiples and implied valuation"),
        ("Sensitivities", dcf_sensitivity, "WACC × terminal-growth implied share price"),
        ("Football Field", football_field, "Valuation range by methodology"),
    ):
        ws = wb[sheet_name]; _title(ws, sheet_name, subtitle, 12)
        if frame is not None and not frame.empty: _write_frame(ws, frame, 4, 1, include_index=True)
        else: ws["A4"] = "Output unavailable"

    ws = wb["Model Checks & Analytics"]
    _title(ws, "Model Checks & Analytics", "Explicit reconciliation controls and historical-versus-forecast analytics", 12)
    row = 4
    if model_checks is not None and not model_checks.empty:
        row = _section(ws, row, "MODEL CHECKS", 12); row, _ = _write_frame(ws, model_checks, row, 1, include_index=False); row += 2
        status_col = next((i + 1 for i, c in enumerate(model_checks.columns) if str(c).lower() == "status"), None)
        if status_col:
            rng = f"{get_column_letter(status_col)}5:{get_column_letter(status_col)}{4+len(model_checks)}"
            ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"PASS"'], fill=PatternFill("solid", fgColor=PASS_FILL)))
            ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"FAIL"'], fill=PatternFill("solid", fgColor=FAIL_FILL)))
    if analytics is not None and not analytics.empty:
        row = _section(ws, row, "HISTORICAL VS FORECAST ANALYTICS", 12); _write_frame(ws, analytics, row, 1, include_index=False)

    # Cover last so links and outputs are known.
    ws = wb["Summary"]
    _title(ws, f"{company_name} ({ticker}) Valuation Model", f"As of {date.today():%B %d, %Y} | $mm except per-share data", 10)
    row = _section(ws, 4, "VALUATION SUMMARY", 6)
    summary = [("Current share price", f"='Assumptions'!B{assumption_cells['Current share price']}", PER_SHARE),
               ("DCF implied share price", f"='DCF'!B{terminal_row+8}", PER_SHARE),
               ("Upside / (downside)", "=B6/B5-1", PERCENT),
               ("WACC", f"='Assumptions'!B{assumption_cells['WACC']}", PERCENT),
               ("Terminal growth", f"='Assumptions'!B{assumption_cells['Terminal growth']}", PERCENT)]
    for r, (label, formula, fmt) in enumerate(summary, row):
        ws.cell(r, 1, label); ws.cell(r, 2, formula); ws.cell(r, 2).number_format = fmt; ws.cell(r, 2).font = Font(name="Arial", color=GREEN, bold=r < 7)
    row = _section(ws, row + len(summary) + 2, "MODEL NAVIGATION", 6)
    for name in names[1:]:
        ws.cell(row, 1, name); ws.cell(row, 1).hyperlink = f"#'{name}'!A1"; ws.cell(row, 1).style = "Hyperlink"; row += 1
    ws["D5"] = "Model status"; ws["E5"] = "PASS" if model_checks is None or model_checks.empty or not model_checks.get("status", pd.Series()).eq("FAIL").any() else "FAIL"
    ws["E5"].fill = PatternFill("solid", fgColor=PASS_FILL if ws["E5"].value == "PASS" else FAIL_FILL)

    for ws in wb.worksheets:
        _polish(ws)
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_margins.left = ws.page_margins.right = 0.25
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(output_path)
    return output_path


def validate_exported_workbook(path, expected_sheets=None):
    """Open the workbook and return structural/formula validation diagnostics."""
    wb = load_workbook(path, data_only=False, read_only=False)
    expected_sheets = expected_sheets or []
    missing = [name for name in expected_sheets if name not in wb.sheetnames]
    formula_errors = []
    formulas = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas += 1
                    if "#REF!" in cell.value:
                        formula_errors.append(f"{ws.title}!{cell.coordinate}")
    return {"opens_cleanly": True, "missing_sheets": missing, "formula_count": formulas,
            "formula_reference_errors": formula_errors, "sheet_count": len(wb.sheetnames)}
